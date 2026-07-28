#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Schmootz exporter.

Reads the Schmootz workbook and writes a small JSON file for schmootz.html.

Why a separate exporter rather than part of oil_scraper.py:
  * Schmootz is displacement between sites, NOT oil collected from customers.
    It must never be mixed into the collection totals.
  * The workbook is updated by hand, not by the nightly scrape, so there is no
    reason to re-read it every night.

The workbook itself is deliberately NOT committed. It is a ~4.6 MB binary that
would store a fresh copy in git history on every re-save, and it contains a lot
of internal working columns the site does not need. Only the generated
schmootz_data.json is committed.

Run:
    python3 export_schmootz.py
"""

import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl is required: pip install openpyxl")


# Local source, in preference order. Both are gitignored.
SOURCE_CANDIDATES = ["data/Schmootz.xlsx", "Schmootz.xlsx"]
SHEET_NAME = "Schmootz"
OUTPUT_PATH = "schmootz_data.json"

MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]

# The two clean blocks in the sheet. Everything else on it (columns V onward,
# rows 39+) is derived scratch working and is deliberately ignored.
#
# Each block is laid out the same way:
#   header row  — "Date", then one column per year
#   rows 4..15  — one row per calendar month
#   row 16      — the sheet's own "Yearly Totals", used only as a cross-check
BLOCKS = [
    {
        "key": "barr_hill_to_gebbie",
        "label": "Barr Hill → Gebbie Farm",
        "source": "Barr Hill",
        "destination": "Gebbie Farm",
        "header_row": 3,
        "first_col": 2,          # column B, the "Date" label column
        "last_col": 9,           # column I, the last year column (2026)
    },
    {
        "key": "bbb_shop_to_gebbie",
        "label": "Black Bear Biodiesel Shop → Gebbie Farm",
        "source": "Black Bear Biodiesel Shop",
        "destination": "Gebbie Farm",
        "header_row": 3,
        "first_col": 13,         # column M
        "last_col": 17,          # column Q
    },
]

MONTH_FIRST_ROW = 4
MONTH_LAST_ROW = 15
TOTALS_ROW = 16


def find_source():
    for path in SOURCE_CANDIDATES:
        if os.path.exists(path):
            return path
    raise SystemExit(
        "Could not find the Schmootz workbook. Looked for:\n  "
        + "\n  ".join(SOURCE_CANDIDATES)
        + "\nPut it at data/Schmootz.xlsx and run this again."
    )


def cell_number(value):
    """
    Workbook cells hold either a number, a blank (month not reached yet), or a
    '-' marker meaning no movement recorded. Both non-numeric cases become None
    so the page can tell 'no data' apart from a real zero.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).strip()
    if text in ("", "-", "—", "N/A", "n/a"):
        return None
    try:
        return int(round(float(text.replace(",", ""))))
    except ValueError:
        return None


def read_block(ws, block):
    """Pull one block into years, a month x year grid, and derived totals."""
    header_row = block["header_row"]

    years = []
    columns = []
    for col in range(block["first_col"] + 1, block["last_col"] + 1):
        raw = ws.cell(row=header_row, column=col).value
        if raw is None:
            continue
        try:
            year = int(float(raw))
        except (TypeError, ValueError):
            continue
        years.append(year)
        columns.append(col)

    if not years:
        raise SystemExit(f"No year columns found for block {block['key']!r}. "
                         "The workbook layout may have changed.")

    monthly = []       # flat {year, month, month_name, gallons}
    by_year = {y: 0 for y in years}
    by_month_name = {name: 0 for name in MONTHS}
    records = 0

    for offset, row in enumerate(range(MONTH_FIRST_ROW, MONTH_LAST_ROW + 1)):
        label = ws.cell(row=row, column=block["first_col"]).value
        month_name = str(label).strip() if label else MONTHS[offset]

        # Guard against a reshuffled sheet rather than silently mis-labelling.
        if month_name not in MONTHS:
            raise SystemExit(
                f"Expected a month name at row {row} of block {block['key']!r}, "
                f"found {month_name!r}. The workbook layout may have changed."
            )
        month_number = MONTHS.index(month_name) + 1

        for year, col in zip(years, columns):
            gallons = cell_number(ws.cell(row=row, column=col).value)
            if gallons is None:
                continue
            monthly.append({
                "year": year,
                "month": f"{year}-{month_number:02d}",
                "month_name": month_name,
                "gallons": gallons,
            })
            by_year[year] += gallons
            by_month_name[month_name] += gallons
            records += 1

    total = sum(by_year.values())

    # The sheet keeps its own "Yearly Totals" row. Compare against it so a
    # layout change or a stale formula is caught here rather than on the page.
    sheet_totals = {}
    for year, col in zip(years, columns):
        value = cell_number(ws.cell(row=TOTALS_ROW, column=col).value)
        if value is not None:
            sheet_totals[year] = value

    mismatches = [
        {"year": y, "computed": by_year[y], "sheet": sheet_totals[y]}
        for y in years
        if y in sheet_totals and sheet_totals[y] != by_year[y]
    ]

    return {
        "key": block["key"],
        "label": block["label"],
        "source": block["source"],
        "destination": block["destination"],
        "years": years,
        "first_year": min(years),
        "last_year": max(years),
        "total_gallons": total,
        "record_count": records,
        "months_with_data": records,
        "by_year": [{"year": y, "gallons": by_year[y]} for y in years],
        "by_month_name": [
            {"month_name": m, "gallons": by_month_name[m]} for m in MONTHS
        ],
        "monthly": monthly,
        "_sheet_total_mismatches": mismatches,
    }


def main():
    source = find_source()
    wb = openpyxl.load_workbook(source, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"{source} has no sheet named {SHEET_NAME!r}. "
                         f"Found: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    print(f"Reading {source} · sheet {SHEET_NAME!r}")
    blocks = [read_block(ws, b) for b in BLOCKS]

    problems = 0
    for block in blocks:
        mismatches = block.pop("_sheet_total_mismatches")
        if mismatches:
            problems += len(mismatches)
            print(f"  WARNING: {block['label']} disagrees with the sheet's own "
                  f"Yearly Totals row:")
            for m in mismatches:
                print(f"    {m['year']}: computed {m['computed']:,} "
                      f"vs sheet {m['sheet']:,}")
        else:
            print(f"  {block['label']}: {block['total_gallons']:>10,} gallons "
                  f"({block['record_count']} months, "
                  f"{block['first_year']}–{block['last_year']}) — "
                  "cross-checks against the sheet's own totals")

    lookup = {b["key"]: b for b in blocks}
    barr = lookup["barr_hill_to_gebbie"]["total_gallons"]
    shop = lookup["bbb_shop_to_gebbie"]["total_gallons"]

    all_years = sorted({y for b in blocks for y in b["years"]})

    # Combined per-year table, ready for the page to render directly.
    combined_by_year = []
    for year in all_years:
        row = {"year": year}
        total = 0
        for b in blocks:
            g = next((e["gallons"] for e in b["by_year"] if e["year"] == year), None)
            row[b["key"]] = g
            total += g or 0
        row["combined"] = total
        combined_by_year.append(row)

    output = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_file": os.path.basename(source),
        "source_sheet": SHEET_NAME,
        "total_barr_hill_to_gebbie": barr,
        "total_bbb_shop_to_gebbie": shop,
        "combined_total": barr + shop,
        "first_year": all_years[0],
        "last_year": all_years[-1],
        "years": all_years,
        "record_count": sum(b["record_count"] for b in blocks),
        "blocks": blocks,
        "combined_by_year": combined_by_year,
    }

    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)

    size_kb = os.path.getsize(OUTPUT_PATH) / 1000
    print(f"\nCombined: {barr + shop:,} gallons "
          f"({all_years[0]}–{all_years[-1]})")
    print(f"Exported {OUTPUT_PATH} ({size_kb:.1f} KB)")
    print("\nThe workbook itself is not committed; only this JSON is.")

    return 1 if problems else 0


if __name__ == "__main__":
    sys.exit(main())
